"""사용자가 지정한 "추출양식.xlsx"와 동일한 구조(일별 근무시간 시트 + 월별 근무시간 시트)로
근무시간 상세 데이터를 엑셀로 뽑아낸다.

시트 레이아웃(2026-08-31 사용자 지정, 바탕화면 `추출양식.xlsx` 참고):
- 일별 근무시간: A~Y, 인원 x 일자 단위로 한 행씩.
- 월별 근무시간: A~L 기본정보 + M~S는 그 달 합산치(근무일수 제외), 인원 x 월 단위로 한 행씩.
- 두 시트 모두 A열("구분")은 "{YY}년 {MM}월" 형식.
- 근무시간류(분 단위로 저장된 값)는 셀 서식을 [h]:mm:ss로 표시해야 해서, 저장값(분)을
  하루 기준 분수(분/1440)로 바꿔서 채워 넣는다.

성능 참고: write_only 모드 + 서식 필요 셀만 WriteOnlyCell로 감싸는 방식도 시도해봤지만
(로컬 18,852행 기준) 오히려 일반 Workbook보다 느려서(67s vs 54s) 되돌렸다 — 서식 적용을
"값 채우기(append, 빠름)"와 "서식 지정(정수 좌표로 순회, 상대적으로 빠름)" 두 단계로 나누는
지금 방식이 더 낫다. 부서 하나(3만행 미만) 기준 실측이니, 상위 N개 부서처럼 훨씬 커지면
체감 대기시간이 늘어날 수 있다는 점은 감안할 것.

주의: 근무스케쥴타입/근무시작/근무종료/당시부서 4개 컬럼은 db.py에 나중에 추가된 컬럼이라,
이 컬럼이 생기기 전에 업로드된 행은 비어 있을 수 있다(재업로드해야 채워짐).
"""

from __future__ import annotations

import io
from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from . import columns as C
from .org import DIVISION_COL

_DURATION_FMT = "[h]:mm:ss"
_DATE_FMT = "yyyy-mm-dd"

_DAILY_HEADERS = [
    "구분", "성명", "사번", "사업부", "부서명", "직급", "일자", "요일",
    "근무스케쥴타입", "근무조", "근무형태", "당시부서", "근무시간", "입문", "출문",
    "★체류시간", "근무시작", "근무종료", "변경후시작", "변경후종료", "근태",
    "제외시간", "연장근무시간", "심야근무시간", "휴일근무시간",
]
_DAILY_DURATION_COLS = {"근무시간", "★체류시간", "제외시간", "연장근무시간", "심야근무시간", "휴일근무시간"}

_MONTHLY_HEADERS = [
    "구분", "성명", "사번", "사업부", "부서명", "직급", "일자", "요일",
    "근무스케쥴타입", "근무조", "근무형태", "당시부서", "근무시간", "★체류시간",
    "근무시작", "제외시간", "연장근무시간", "심야근무시간", "휴일근무시간",
]
_MONTHLY_DURATION_COLS = {"근무시간", "★체류시간", "제외시간", "연장근무시간", "심야근무시간", "휴일근무시간"}

_MINUTE_COL = {
    "근무시간": "worktime_minutes",
    "★체류시간": "stay_minutes",
    "제외시간": "exclude_minutes",
    "연장근무시간": "overtime_minutes",
    "심야근무시간": "night_minutes",
    "휴일근무시간": "holiday_work_minutes",
}


def _gubun_label(yyyymmdd: str) -> str:
    """'20260731' -> '26년 07월'."""
    return f"{yyyymmdd[2:4]}년 {yyyymmdd[4:6]}월"


def _to_date(yyyymmdd: str) -> "date | None":
    try:
        return date(int(yyyymmdd[:4]), int(yyyymmdd[4:6]), int(yyyymmdd[6:8]))
    except (ValueError, TypeError):
        return None


def _minutes_to_day_fraction(minutes) -> float:
    if minutes is None or pd.isna(minutes):
        return 0.0
    return float(minutes) / 1440.0


def _write_header(ws, headers: "list[str]") -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, len(h) * 1.6)


def _apply_number_format(ws, col_idx: int, fmt: str, n_rows: int) -> None:
    # 문자열 좌표("A2")로 셀을 찾으면 매번 좌표 파싱이 들어가서 느리다 — 정수 좌표(row/column)로
    # 바로 접근하면 더 빠르다.
    for r in range(2, n_rows + 2):
        ws.cell(row=r, column=col_idx).number_format = fmt


def _apply_duration_formats(ws, headers: "list[str]", duration_cols: set, n_rows: int) -> None:
    for i, h in enumerate(headers, start=1):
        if h in duration_cols:
            _apply_number_format(ws, i, _DURATION_FMT, n_rows)


def _write_daily_sheet(ws, df: pd.DataFrame) -> int:
    _write_header(ws, _DAILY_HEADERS)
    ordered = df.sort_values([C.COL_EMP_ID, C.COL_DATE])
    n = 0
    # itertuples()/_asdict()는 "★체류시간"처럼 유효한 파이썬 식별자가 아닌 컬럼명을 자동으로
    # _1/_2 같은 위치기반 이름으로 바꿔버려서(rename=True 기본동작) 원래 이름으로 조회가 안 된다.
    # to_dict(orient="records")는 원래 컬럼명 문자열을 그대로 키로 쓰므로 이 문제가 없다.
    for r in ordered.to_dict(orient="records"):
        ws.append(
            [
                _gubun_label(r[C.COL_DATE]),
                r[C.COL_NAME],
                r[C.COL_EMP_ID],
                r.get(DIVISION_COL),
                r[C.COL_DEPT],
                r[C.COL_RANK],
                _to_date(r[C.COL_DATE]),
                r[C.COL_DOW],
                r.get(C.COL_SCHEDULE_TYPE) or "",
                r.get(C.COL_WORKGROUP),
                r.get(C.COL_WORKTYPE),
                r.get(C.COL_CUR_DEPT) or r[C.COL_DEPT],
                _minutes_to_day_fraction(r.get("worktime_minutes")),
                r.get(C.COL_CHECKIN),
                r.get(C.COL_CHECKOUT),
                _minutes_to_day_fraction(r.get("stay_minutes")),
                r.get(C.COL_WORK_START) or "",
                r.get(C.COL_WORK_END) or "",
                r.get(C.COL_CHANGED_START),
                r.get(C.COL_CHANGED_END),
                r.get(C.COL_ATTENDANCE),
                _minutes_to_day_fraction(r.get("exclude_minutes")),
                _minutes_to_day_fraction(r.get("overtime_minutes")),
                _minutes_to_day_fraction(r.get("night_minutes")),
                _minutes_to_day_fraction(r.get("holiday_work_minutes")),
            ]
        )
        n += 1
    g_col = _DAILY_HEADERS.index("일자") + 1
    _apply_number_format(ws, g_col, _DATE_FMT, n)
    _apply_duration_formats(ws, _DAILY_HEADERS, _DAILY_DURATION_COLS, n)
    return n


def _write_monthly_sheet(ws, df: pd.DataFrame) -> int:
    _write_header(ws, _MONTHLY_HEADERS)
    d = df.copy()
    d["_month"] = d[C.COL_DATE].str[:6]
    # daily_records엔 근무 여부와 무관하게 달력상 매일 한 행씩 있어서, 단순 날짜 개수(nunique)는
    # "이번 달 며칠까지 데이터가 쌓였는지"일 뿐 실제 근무일수가 아니다. 실제로 일한 날만 센다.
    d["_worked_day"] = d["worktime_minutes"].fillna(0) > 0

    agg_spec = {col: (minute_col, "sum") for col, minute_col in _MINUTE_COL.items()}
    agg = (
        d.groupby([C.COL_EMP_ID, "_month"])
        .agg(
            **agg_spec,
            근무일수=("_worked_day", "sum"),
            _profile_row=(C.COL_DATE, "idxmax"),
        )
        .reset_index()
    )
    profile_cols = [
        C.COL_NAME, DIVISION_COL, C.COL_DEPT, C.COL_RANK,
        C.COL_SCHEDULE_TYPE, C.COL_WORKGROUP, C.COL_WORKTYPE, C.COL_CUR_DEPT,
    ]
    profile = d.loc[agg["_profile_row"], [c for c in profile_cols if c in d.columns]].reset_index(drop=True)
    agg = pd.concat([agg.drop(columns=["_profile_row"]), profile], axis=1)
    agg = agg.sort_values([C.COL_EMP_ID, "_month"])

    n = 0
    for r in agg.to_dict(orient="records"):
        month = r["_month"]
        ws.append(
            [
                f"{month[2:4]}년 {month[4:6]}월",
                r[C.COL_NAME],
                r[C.COL_EMP_ID],
                r.get(DIVISION_COL),
                r[C.COL_DEPT],
                r[C.COL_RANK],
                None,  # 일자: 월 집계라 특정 일자가 없음
                None,  # 요일: 월 집계라 없음
                r.get(C.COL_SCHEDULE_TYPE) or "",
                r.get(C.COL_WORKGROUP),
                r.get(C.COL_WORKTYPE),
                r.get(C.COL_CUR_DEPT) or r[C.COL_DEPT],
                _minutes_to_day_fraction(r.get("근무시간")),
                _minutes_to_day_fraction(r.get("★체류시간")),
                int(r["근무일수"]),
                _minutes_to_day_fraction(r.get("제외시간")),
                _minutes_to_day_fraction(r.get("연장근무시간")),
                _minutes_to_day_fraction(r.get("심야근무시간")),
                _minutes_to_day_fraction(r.get("휴일근무시간")),
            ]
        )
        n += 1
    _apply_duration_formats(ws, _MONTHLY_HEADERS, _MONTHLY_DURATION_COLS, n)
    return n


def build_worktime_workbook(df: pd.DataFrame) -> "tuple[io.BytesIO, int, int]":
    """scoped daily_records df(사업부 포함)로 (일별 근무시간 + 월별 근무시간) 워크북을 만든다.

    반환: (엑셀 바이트 버퍼, 일별 시트 행 수, 월별 시트 행 수)
    """
    wb = Workbook()
    daily_ws = wb.active
    daily_ws.title = "일별 근무시간"
    daily_rows = _write_daily_sheet(daily_ws, df)

    monthly_ws = wb.create_sheet("월별 근무시간")
    monthly_rows = _write_monthly_sheet(monthly_ws, df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, daily_rows, monthly_rows
